"""Validate build_rekon() against Salsa's real April Palembang REKON sheet."""
import warnings
import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")
from engine import (Config, read_sap, read_coretax, read_etb, build_rekon,
                    build_cabang_index)

BASE = "../WORK MAY 2026/"
SAP = BASE + "PPN PALEMBANG/21 APR - 20 MEI PPN WAPU SAP RO PALEMBANG APRIL 2026.xlsx"
CORETAX = BASE + "PPN PALEMBANG/EKSPOR PSIAP PPN WAPU PALEMBANG APRIL 2026.xlsx"
ETB = BASE + "ETB PPN WAPU - 20 Mei 2026.xlsx"

cfg = Config(ro_name="PALEMBANG", masa=4, tahun=2026)
sap = read_sap(SAP)
coretax = read_coretax(CORETAX, "Faktur Masukan_1")
etb = read_etb(ETB, ro_name="PALEMBANG")
cbf, cba, cbs = build_cabang_index(SAP)
rekon, flagged = build_rekon(coretax, sap, etb, cfg, cbf, cba, cbs)

# real REKON
wb = openpyxl.load_workbook(SAP, data_only=True)
rk = wb["REKON"]
hdr = [c.value for c in rk[1]]
real = {}
for i in range(2, rk.max_row + 1):
    u = rk.cell(i, 1).value
    if u in (None, ""):
        continue
    months = {hdr[j]: rk.cell(i, j + 1).value for j in range(1, 17) if rk.cell(i, j + 1).value not in (None, 0)}
    real[int(u)] = {"months": months, "ETB": rk.cell(i, 19).value, "SELISIH": rk.cell(i, 20).value}
wb.close()

gen = {int(r["KODE UKER"]): r for _, r in rekon.iterrows()}
print(f"Generated rekon: {len(gen)} ukers | real: {len(real)} ukers | flagged: {len(flagged)}")

mm_month = mm_etb = mm_sel = ok_month = ok_etb = ok_sel = 0
for u, info in real.items():
    g = gen.get(u)
    if g is None:
        print(f"  uker {u} MISSING from generated"); continue
    for mlabel, val in info["months"].items():
        gv = g.get(mlabel)
        if gv is not None and abs(gv - val) < 1: ok_month += 1
        else: mm_month += 1; print(f"  month diff uker {u} {mlabel}: real={val:,.0f} gen={gv}")
    if info["ETB"] is not None:
        if g.get("ETB") is not None and abs(g["ETB"] - info["ETB"]) < 1: ok_etb += 1
        else: mm_etb += 1
    if info["SELISIH"] is not None:
        if g.get("SELISIH") is not None and abs(g["SELISIH"] - info["SELISIH"]) < 1: ok_sel += 1
        else: mm_sel += 1

print(f"\nMonths : {ok_month} ok, {mm_month} mismatch")
print(f"ETB    : {ok_etb} ok, {mm_etb} mismatch")
print(f"SELISIH: {ok_sel} ok, {mm_sel} mismatch")
if len(flagged):
    print(f"\nFlagged (reclass / unmapped):\n{flagged.to_string(index=False)}")
