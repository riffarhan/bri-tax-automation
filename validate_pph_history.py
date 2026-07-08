"""
PPh stress-test across EVERY month with complete data: masa Des 2025 ... Mei 2026
(WORK JANUARY ... WORK JUNE), both ROs, all 4 streams + DIO manual pairs.
File naming varies wildly per month, so discovery is keyword-based.
"""
import glob
import os
import re
import warnings

import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")
from engine_pph import (PphConfig, read_sap_pph, read_sipo, read_dio_pph,
                        build_template_sap, build_template_sipo, build_template_dio,
                        read_etb_pph, build_rekon_pph, _uker_from_label)

ROOT = ".."
RUNS = [
    ("WORK JANUARY 2026", 12, 2025, "DESEMBER"),
    ("WORK FEBRUARY 2026", 1, 2026, "JANUARI"),
    ("WORK MARCH 2026", 2, 2026, "FEBRUARI"),
    ("WORK APRIL 2026", 3, 2026, "MARET"),
    ("WORK MAY 2026", 4, 2026, "APRIL"),
    ("WORK JUNE 2026", 5, 2026, "MEI"),
]
PASAL_TOKENS = {"22": ["PPH 22"], "23": ["PPH 23"],
                "4A2": ["PPH 4 AYAT 2", "PPH 4-2", "4 AYAT 2", "PPH 4 "]}
BAD_TOKENS = ["KETINGGALAN", "NOVEMBER"]  # stray prior-masa files

CRITICAL = ["NPWP Pemotong", "NITKU Pemotong (6 Digit Terakhir)", "Masa Pajak",
            "Tahun Pajak", "NPWP Penerima Penghasilan",
            "NITKU Penerima Penghasilan (22 Digit)", "Jenis PPh",
            "Kode Objek Pajak", "Fasilitas Insentif", "Nomor Setifikat Insentif",
            "Tarif Fasilitas", "Penghasilan Bruto", "Jenis Dokumen Referensi",
            "NPWP Penandatangan"]


def ls(folder, ro):
    return [f for f in glob.glob(f"{ROOT}/{folder}/PPH {ro}/*.xls*") if "~$" not in f]


def has_pasal(name, pasal):
    up = name.upper()
    if any(t in up for t in PASAL_TOKENS[pasal]):
        # "PPH 4 AYAT 2" contains neither PPH 22 nor 23 tokens; but "PPH 23"
        # matches inside "PPH 23 MANUAL" etc. — pasal 22/23 must not be 4A2 file
        if pasal in ("22", "23") and any(t in up for t in PASAL_TOKENS["4A2"]):
            return False
        return True
    return False


def sap_file(folder, ro, pasal):
    cands = [f for f in ls(folder, ro)
             if has_pasal(os.path.basename(f), pasal)
             and ("SAP" in os.path.basename(f).upper())
             and not any(t in os.path.basename(f).upper()
                         for t in ["TEMPLATE", "PSIAP", "EKSPOR", "REVISI"] + BAD_TOKENS)]
    return sorted(cands, key=lambda f: len(os.path.basename(f)))[0] if cands else None


def tpl_file(folder, ro, pasal_or_sipo, manual=False):
    base, fallback = [], []
    for f in ls(folder, ro):
        b = os.path.basename(f).upper()
        if "NEW TEMPLATE" not in b or "PSIAP" not in b:
            continue
        if ("MANUAL" in b) != manual:
            continue
        if pasal_or_sipo == "SIPOBRI":
            if "SIPOBRI" not in b.replace(" ", ""):
                continue
        elif "SIPOBRI" in b.replace(" ", "") or not has_pasal(b, pasal_or_sipo):
            continue
        if any(t in b for t in ["SISA", "REVISI", "UPLOAD KE", "KEDUA", "KETIGA"]):
            fallback.append(f)
        else:
            base.append(f)
    pick = base or fallback
    return sorted(pick, key=lambda f: len(os.path.basename(f)))[0] if pick else None


def sipo_dir(folder, ro):
    for d in glob.glob(f"{ROOT}/{folder}/PPH {ro}/SIPO*"):
        if os.path.isdir(d):
            files = [f for f in glob.glob(d + "/*.xls") if "~$" not in f]
            if files:
                return files
    return []


def dio_file(folder, ro):
    for d in glob.glob(f"{ROOT}/{folder}/PPH {ro}/*") + glob.glob(f"{ROOT}/{folder}/*"):
        if os.path.isdir(d) and ("SENDIK" in os.path.basename(d).upper()
                                 or "MANUAL" in os.path.basename(d).upper()):
            for f in glob.glob(d + "/*.xlsx"):
                if "PPH" in os.path.basename(f).upper() and "~$" not in f:
                    return f
    return None


def etb_file(folder, ro):
    cands = [f for f in glob.glob(f"{ROOT}/{folder}/PPH {ro}/*ETB*.xls*")
             + glob.glob(f"{ROOT}/{folder}/*ETB*PPh*.xls*")
             + glob.glob(f"{ROOT}/{folder}/*ETB*Unifikasi*.xls*") if "~$" not in f]
    return cands[0] if cands else None


def norm_cell(c, v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if c == "Penghasilan Bruto":
        try:
            return str(round(float(v)))
        except (TypeError, ValueError):
            return str(v).strip()
    if "Tanggal" in c:
        try:
            return pd.to_datetime(v).strftime("%Y-%m-%d")
        except Exception:
            return str(v).strip()
    return str(v).strip()


def compare(gen, tpl_path, label):
    exp = pd.read_excel(tpl_path, sheet_name="Template", dtype=object)
    exp.columns = [str(c).strip() for c in exp.columns]
    exp = exp[exp[exp.columns[0]].notna()].reset_index(drop=True)

    def key(df):
        return list(zip(df["Penghasilan Bruto"].map(lambda v: norm_cell("Penghasilan Bruto", v)),
                        df["Nomor Dokumen Referensi"].map(lambda v: str(v).strip()),
                        df["NITKU Pemotong (6 Digit Terakhir)"].map(lambda v: str(v).strip())))
    gk, ek = key(gen), key(exp)
    used, pairs = set(), []
    for tier in (lambda a, b: a == b, lambda a, b: a[:2] == b[:2],
                 lambda a, b: (a[0], a[2]) == (b[0], b[2]), lambda a, b: a[0] == b[0]):
        for gi, k in enumerate(gk):
            if gi in {p[0] for p in pairs}:
                continue
            for ei, k2 in enumerate(ek):
                if ei not in used and tier(k, k2):
                    pairs.append((gi, ei)); used.add(ei); break
    tot = mm = 0
    bycol = {}
    for gi, ei in pairs:
        for c in CRITICAL:
            if c not in exp.columns:
                continue
            tot += 1
            a, b = norm_cell(c, gen.iloc[gi][c]), norm_cell(c, exp.iloc[ei][c])
            if a != b:
                mm += 1
                bycol[c] = bycol.get(c, 0) + 1
    pct = 100 * (tot - mm) / tot if tot else 0
    worst = max(bycol, key=bycol.get) if bycol else ""
    return (f"rows {len(gen)}/{len(exp)} matched {len(pairs)} | KRITIS {pct:.1f}%"
            + (f" (terbanyak: {worst} {bycol[worst]}x)" if worst else ""))


def rekon_check(sap_path, res, folder, ro, pasal):
    etbf = etb_file(folder, ro)
    if not etbf:
        return "ETB tidak ada"
    utang = read_etb_pph(etbf, ro, pasal)
    if not utang:
        return f"ETB kosong ({os.path.basename(etbf)[:30]})"
    rekon = build_rekon_pph(res.recon_rows, utang, ro)
    gen = {int(r["KODE UKER"]): r for _, r in rekon.iterrows()}
    wb = openpyxl.load_workbook(sap_path, read_only=True, data_only=True)
    s2 = next((s for s in wb.sheetnames if s.strip().upper() == "SHEET2"), None)
    if not s2:
        wb.close(); return f"utang {len(utang)} ukers (tidak ada Sheet2)"
    rows = list(wb[s2].iter_rows(values_only=True))
    wb.close()
    hi = next((i for i, r in enumerate(rows)
               if r and any("PCA L2" in str(c) for c in r if c)), None)
    if hi is None:
        return f"utang {len(utang)} ukers (Sheet2 tanpa header)"
    rows = rows[hi:]
    hdr = [str(c).strip().upper() if c else "" for c in rows[0]]
    pj_i = hdr.index("TOTAL") if "TOTAL" in hdr else (hdr.index("PAJAK") if "PAJAK" in hdr else None)
    sel_i = hdr.index("SELISIH") if "SELISIH" in hdr else None
    oks = ms = 0
    for r in rows[1:]:
        if not r or r[0] is None or "GRAND" in str(r[0]).upper():
            continue
        uker = _uker_from_label(r[0], ro)
        if uker is None or sel_i is None or sel_i >= len(r) or r[sel_i] is None:
            continue
        g = gen.get(uker)
        try:
            gs = g["SELISIH"] if g is not None else None
            oks, ms = (oks + 1, ms) if (gs is not None and abs(gs - float(r[sel_i])) < 2) else (oks, ms + 1)
        except (TypeError, ValueError):
            pass
    return f"SELISIH {oks}ok/{ms}mm"


print("=" * 96)
print("VALIDASI HISTORIS PPh — Des 2025 s/d Mei 2026")
print("=" * 96)
for folder, masa, tahun, bln in RUNS:
    for ro in ["PALEMBANG", "YOGYAKARTA"]:
        cfg = PphConfig(ro_name=ro, masa=masa, tahun=tahun)
        print(f"\n### {bln} {tahun} — {ro} ({folder})")
        for pasal in ["22", "23", "4A2"]:
            sapf, tplf = sap_file(folder, ro, pasal), tpl_file(folder, ro, pasal)
            if not sapf or not tplf:
                print(f"  PPh {pasal:4}: SKIP (sap={bool(sapf)} tpl={bool(tplf)})")
                continue
            try:
                sap = read_sap_pph(sapf)
                if sap.empty:
                    print(f"  PPh {pasal:4}: SAP kosong ({os.path.basename(sapf)[:40]})")
                    continue
                res = build_template_sap(sap, cfg, pasal)
                print(f"  PPh {pasal:4}: {compare(res.template, tplf, pasal)}"
                      f" | rekon {rekon_check(sapf, res, folder, ro, pasal)}")
            except Exception as e:
                print(f"  PPh {pasal:4}: ❌ {type(e).__name__}: {e}")
        sfiles, stpl = sipo_dir(folder, ro), tpl_file(folder, ro, "SIPOBRI")
        if sfiles and stpl:
            try:
                res = build_template_sipo(read_sipo(sfiles), cfg)
                print(f"  SIPOBRI : {compare(res.template, stpl, 'SIPOBRI')} ({len(sfiles)} file)")
            except Exception as e:
                print(f"  SIPOBRI : ❌ {type(e).__name__}: {e}")
        else:
            print(f"  SIPOBRI : SKIP (files={len(sfiles)} tpl={bool(stpl)})")
        diof, mtpl = dio_file(folder, ro), tpl_file(folder, ro, "23", manual=True)
        if diof and mtpl:
            try:
                res = build_template_dio(read_dio_pph(diof), cfg, "23")
                print(f"  MANUAL  : {compare(res.template, mtpl, 'DIO')}")
            except Exception as e:
                print(f"  MANUAL  : ❌ {type(e).__name__}: {e}")
print("\n" + "=" * 96)
