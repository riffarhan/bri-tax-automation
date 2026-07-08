"""
Validate the PPh template generators against Salsa's real NEW TEMPLATE PSIAP
files — April (WORK MAY 2026) + Maret (WORK APRIL 2026), Palembang & Yogyakarta.
Rows matched by (Penghasilan Bruto, Nomor Dokumen Referensi) then bruto alone.
"""
import glob
import os
import warnings

import pandas as pd

warnings.filterwarnings("ignore")
from engine_pph import (PphConfig, read_sap_pph, read_sipo,
                        build_template_sap, build_template_sipo,
                        TEMPLATE_COLUMNS)

ROOT = ".."
RUNS = [("WORK APRIL 2026", 3, 2026, "MARET"), ("WORK MAY 2026", 4, 2026, "APRIL")]
PASAL_FILES = {"22": "PPH 22", "23": "PPH 23", "4A2": "PPH 4 AYAT 2"}

# columns whose values must be right for the upload to be correct
CRITICAL = ["NPWP Pemotong", "NITKU Pemotong (6 Digit Terakhir)", "Masa Pajak",
            "Tahun Pajak", "NPWP Penerima Penghasilan",
            "NITKU Penerima Penghasilan (22 Digit)", "Jenis PPh",
            "Kode Objek Pajak", "Fasilitas Insentif", "Nomor Setifikat Insentif",
            "Tarif Fasilitas", "Penghasilan Bruto", "Jenis Dokumen Referensi",
            "NPWP Penandatangan"]
SOFT = ["Nama Penerima Penghasilan", "Nomor Dokumen Referensi",
        "Tanggal Dokumen Referensi", "Tanggal Pemotongan", "Referensi"]


def find(folder, ro, must, nots=()):
    out = []
    for f in glob.glob(f"{ROOT}/{folder}/PPH {ro}/*.xls*"):
        b = os.path.basename(f).upper()
        if "~$" in b:
            continue
        if all(m in b for m in must) and not any(n in b for n in nots):
            out.append(f)
    return sorted(out)


def norm_cell(c, v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if c in ("Penghasilan Bruto",):
        try:
            return str(round(float(v)))
        except (TypeError, ValueError):
            return str(v).strip()
    if "Tanggal" in c:
        try:
            return pd.to_datetime(v).strftime("%Y-%m-%d")
        except Exception:
            return str(v).strip()
    return str(v).strip()


def compare(gen: pd.DataFrame, tpl_path: str, label: str):
    exp = pd.read_excel(tpl_path, sheet_name="Template", dtype=object)
    exp.columns = [str(c).strip() for c in exp.columns]
    exp = exp[exp[exp.columns[0]].notna()].reset_index(drop=True)

    def key(df):
        return list(zip(df["Penghasilan Bruto"].map(lambda v: norm_cell("Penghasilan Bruto", v)),
                        df["Nomor Dokumen Referensi"].map(lambda v: str(v).strip()),
                        df["NITKU Pemotong (6 Digit Terakhir)"].map(lambda v: str(v).strip())))
    gk, ek = key(gen), key(exp)
    used = set()
    pairs = []
    # tier 1: bruto + no dok + nitku (exact row), tier 2: bruto + no dok, tier 3: bruto
    for tier in (lambda a, b: a == b,
                 lambda a, b: a[:2] == b[:2],
                 lambda a, b: a[0] == b[0]):
        for gi, k in enumerate(gk):
            if gi in {p[0] for p in pairs}:
                continue
            for ei, k2 in enumerate(ek):
                if ei not in used and tier(k, k2):
                    pairs.append((gi, ei)); used.add(ei); break

    crit_tot = crit_mm = soft_tot = soft_mm = 0
    examples = []
    for gi, ei in pairs:
        for c in CRITICAL:
            if c not in exp.columns:
                continue
            a, b = norm_cell(c, gen.iloc[gi][c]), norm_cell(c, exp.iloc[ei][c])
            crit_tot += 1
            if a != b:
                crit_mm += 1
                if len(examples) < 4:
                    examples.append(f"{c}: kita='{a[:24]}' real='{b[:24]}'")
        for c in SOFT:
            if c not in exp.columns:
                continue
            a, b = norm_cell(c, gen.iloc[gi][c]), norm_cell(c, exp.iloc[ei][c])
            soft_tot += 1
            if a != b:
                soft_mm += 1
    unmatched_g = len(gen) - len(pairs)
    unmatched_e = len(exp) - len(pairs)
    pct = 100 * (crit_tot - crit_mm) / crit_tot if crit_tot else 0
    spct = 100 * (soft_tot - soft_mm) / soft_tot if soft_tot else 0
    print(f"  {label}: rows kita={len(gen)} real={len(exp)} matched={len(pairs)} "
          f"(+{unmatched_g}/-{unmatched_e}) | KRITIS {pct:.1f}% | soft {spct:.1f}%")
    for e in examples:
        print(f"      diff {e}")


def rekon_check(sap_path, res, folder, ro, pasal, cfg):
    """Compare our per-uker PAJAK/SELISIH against Salsa's Sheet2."""
    import openpyxl
    from engine_pph import read_etb_pph, build_rekon_pph, _uker_from_label
    etbs = [f for f in glob.glob(f"{ROOT}/{folder}/*ETB PPh*.xls*")
            + glob.glob(f"{ROOT}/{folder}/*ETB*Unifikasi*.xls*") if "~$" not in f]
    if not etbs:
        return "(ETB tidak ada)"
    utang = read_etb_pph(etbs[0], ro, pasal)
    rekon = build_rekon_pph(res.recon_rows, utang, ro)
    gen = {int(r["KODE UKER"]): r for _, r in rekon.iterrows()}
    wb = openpyxl.load_workbook(sap_path, read_only=True, data_only=True)
    if "Sheet2" not in wb.sheetnames:
        wb.close(); return "(tidak ada Sheet2)"
    ws = wb["Sheet2"]
    rows = list(ws.iter_rows(values_only=True))
    hi = next((i for i, r in enumerate(rows)
               if r and any("PCA L2" in str(c) for c in r if c)), 0)
    rows = rows[hi:]
    hdr = [str(c).strip().upper() if c else "" for c in rows[0]]
    # her per-row tax total: TOTAL when the pasal splits buckets (4A2), else PAJAK
    pj_i = hdr.index("TOTAL") if "TOTAL" in hdr else (hdr.index("PAJAK") if "PAJAK" in hdr else None)
    sel_i = hdr.index("SELISIH") if "SELISIH" in hdr else None
    okp = mp = oks = ms = 0
    for r in rows[1:]:
        if not r or r[0] is None or "GRAND" in str(r[0]).upper():
            continue
        uker = _uker_from_label(r[0], ro)
        if uker is None:
            continue
        g = gen.get(uker)
        gp = g["PAJAK"] if g is not None and g["PAJAK"] is not None else 0
        gs = g["SELISIH"] if g is not None else None
        if pj_i is not None and r[pj_i] is not None and str(r[pj_i]).strip() != "":
            try:
                okp, mp = (okp + 1, mp) if abs(gp - float(r[pj_i])) < 2 else (okp, mp + 1)
            except (TypeError, ValueError):
                pass
        if sel_i is not None and r[sel_i] is not None and str(r[sel_i]).strip() != "":
            try:
                oks, ms = (oks + 1, ms) if (gs is not None and abs(gs - float(r[sel_i])) < 2) else (oks, ms + 1)
            except (TypeError, ValueError):
                pass
    wb.close()
    return f"PAJAK {okp}ok/{mp}mm SELISIH {oks}ok/{ms}mm"


print("=" * 92)
print("VALIDASI PPh — generator vs template real")
print("=" * 92)
for folder, masa, tahun, bln in RUNS:
    for ro in ["PALEMBANG", "YOGYAKARTA"]:
        cfg = PphConfig(ro_name=ro, masa=masa, tahun=tahun)
        print(f"\n### {folder} / {ro} (masa {masa:02d}/{tahun})")
        for pasal, fpat in PASAL_FILES.items():
            saps = find(folder, ro, [fpat, "SAP"])
            tpls = find(folder, ro, ["NEW TEMPLATE", fpat.replace("PPH ", "PPH "),
                                     "PSIAP"], nots=["SIPOBRI", "MANUAL"])
            tpls = [t for t in tpls if fpat in os.path.basename(t).upper()]
            if not saps or not tpls:
                print(f"  {pasal}: SKIP (sap={len(saps)} tpl={len(tpls)})")
                continue
            try:
                sap = read_sap_pph(saps[0])
                if sap.empty:
                    print(f"  {pasal}: SAP kosong ({os.path.basename(saps[0])})")
                    continue
                res = build_template_sap(sap, cfg, pasal)
                compare(res.template, tpls[0], f"PPh {pasal} ({res.stats['rows']} baris, "
                        f"{res.stats['exceptions']} exc)")
                print(f"      rekon: {rekon_check(saps[0], res, folder, ro, pasal, cfg)}")
            except Exception as e:
                print(f"  {pasal}: ❌ {type(e).__name__}: {e}")
        # SIPOBRI
        sipo_files = [f for f in glob.glob(f"{ROOT}/{folder}/PPH {ro}/SIPO {ro}/*.xls")
                      if "~$" not in f]
        stpl = find(folder, ro, ["SIPOBRI"])
        if sipo_files and stpl:
            try:
                sipo = read_sipo(sipo_files)
                res = build_template_sipo(sipo, cfg)
                compare(res.template, stpl[0], f"SIPOBRI ({res.stats['rows']} baris, "
                        f"{res.stats['exceptions']} exc, {len(sipo_files)} file)")
            except Exception as e:
                print(f"  SIPOBRI: ❌ {type(e).__name__}: {e}")
        else:
            print(f"  SIPOBRI: SKIP (files={len(sipo_files)} tpl={len(stpl)})")
print("\n" + "=" * 92)
