"""
PPh Unifikasi engine — bukti potong templates for PSIAP.

Four streams share one 26-column Template layout (verified identical across the
real April/Maret files): PPh 22 / 23 / 4 ayat 2 come from the SAP pull, SIPOBRI
comes from the BRITAX per-uker .xls exports. Mapping decoded from
WORK MAY 2026 + WORK APRIL 2026 PPH PALEMBANG (see PPH_ANALYSIS.md):

  NPWP Pemotong        = BRI HO constant (same as PPN)
  NITKU Pemotong       = last 6 of SAP's own NITKU column (master fallback)
  NPWP Penerima        = SAP NIK/NPWP/TIN; all-zeros -> left for Salsa (exception)
  NITKU Penerima       = NPWP penerima + "000000" (22 digits)
  Kode Objek Pajak     = SAP KOP verbatim
  Penghasilan Bruto    = SAP DPP WHT
  Nomor Dok Referensi  = SAP Nomor Invoice; zeros/blank -> Dokumen Invoice
  Tanggal Dok Referensi= SAP Tanggal Invoice (Maret 18/18; Salsa may override)
  Tanggal Pemotongan   = SAP Tanggal Pembayaran
  Referensi            = {JENIS}{SRC}RO{RO}{BULAN}{TAHUN}_{Dokumen Invoice}
  constants            : Fasilitas Insentif=9, Nomor Setifikat=-, Tarif Fasilitas=0,
                         Jenis Dokumen Referensi=02, NPWP Penandatangan (signer)
"""
import io
import os
import re
from dataclasses import dataclass, field

import pandas as pd

from engine import (BULAN_ID, norm_id, to_num, to_date, _load_csv_map,
                    CABANG_INDUK, UKER_MASTER, UKER_NITKU)

TEMPLATE_COLUMNS = [
    "NPWP Pemotong", "NITKU Pemotong (6 Digit Terakhir)", "Masa Pajak",
    "Tahun Pajak", "NPWP Penerima Penghasilan",
    "NITKU Penerima Penghasilan (22 Digit)", "Nama Penerima Penghasilan",
    "Email", "Jenis PPh", "Kode Objek Pajak", "Fasilitas Insentif",
    "Nomor Setifikat Insentif", "Tarif Fasilitas", "Penghasilan Bruto",
    "Jenis Dokumen Referensi", "Nomor Dokumen Referensi",
    "Tanggal Dokumen Referensi",
    "Metode Pembayaran bagi Pemotong Instansi Pemerintah", "Nomor SP2D",
    "NPWP Penandatangan", "Tanggal Pemotongan", "User Id", "Referensi",
    "Referensi 3", "Referensi 4", "Referensi 5",
]

# jenis label in the template + the tag prefix used in the Referensi column
PASAL = {
    "22":      {"jenis": "PPH22",  "tag": "PPH22SAP"},
    "23":      {"jenis": "PPH23",  "tag": "PPH23SAP"},
    "4A2":     {"jenis": "PPH4-2", "tag": "PPH4A2SAP"},
    "SIPOBRI": {"jenis": "PPH23",  "tag": "PPH23SIPO"},
}

BULAN_LONG = {"JANUARI": 1, "FEBRUARI": 2, "MARET": 3, "APRIL": 4, "MEI": 5,
              "JUNI": 6, "JULI": 7, "AGUSTUS": 8, "SEPTEMBER": 9,
              "OKTOBER": 10, "NOVEMBER": 11, "DESEMBER": 12,
              "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "MAY": 5, "JUNE": 6,
              "JULY": 7, "AUGUST": 8, "OCTOBER": 10, "DECEMBER": 12}


@dataclass
class PphConfig:
    ro_name: str = "PALEMBANG"
    masa: int = 4
    tahun: int = 2026
    npwp_pemotong: str = "0010016087093000"       # BRI HO
    npwp_penandatangan: str = "3520116005800002"  # signer (verified constant)

    def tag(self, pasal: str) -> str:
        return f"{PASAL[pasal]['tag']}RO{self.ro_name}{BULAN_ID[self.masa]}{self.tahun}"


@dataclass
class PphResult:
    template: pd.DataFrame
    exceptions: pd.DataFrame
    recon_rows: pd.DataFrame = None     # per-row (uker, pajak) for the ETB rekon
    stats: dict = field(default_factory=dict)


def _zeros(v) -> bool:
    s = norm_id(v)
    return (not s) or set(s) == {"0"}


def _npwp16(v) -> str:
    """Any NPWP/NIK spelling -> bare digits; old 15-digit gets the leading 0
    (SAP Yogya carries formatted values like '01.354.167.7-043.000'). A 12-digit
    value is an old NPWP missing its '000' branch suffix."""
    digits = re.sub(r"\D", "", str(v or ""))
    if len(digits) == 12:
        digits += "000"
    if len(digits) == 15:
        return "0" + digits
    return digits


def _nitku_from_uker(uker) -> str:
    """last-6 NITKU for a uker code, folding branch->induk if needed."""
    if uker is None:
        return ""
    try:
        u = int(float(str(uker).strip()))
    except (TypeError, ValueError):
        return ""
    u = CABANG_INDUK.get(u, u)
    return UKER_NITKU.get(str(u), "") or UKER_NITKU.get(u, "")


# ---------------------------------------------------------------- SAP reader

SAP_PPH_MARKERS = ("Kode Cabang Transaksi", "KOP")


def _is_pph_pull_sheet(header_cells) -> bool:
    vals = {str(c).strip() for c in header_cells if c is not None}
    return all(m in vals for m in SAP_PPH_MARKERS)


def read_sap_pph(path_or_buf) -> pd.DataFrame:
    """Concat every SAP pull sheet (RO / KANINS / SENDIK pulls are identical in
    format per Salsa — 'tetap aku impor bareng'). DATA OLAH (header on row 2,
    Salsa's output) and Sheet2 (rekon) are skipped."""
    import openpyxl
    if hasattr(path_or_buf, "seek"):
        path_or_buf.seek(0)
    wb = openpyxl.load_workbook(path_or_buf, read_only=True, data_only=True)
    frames = []
    for sn in wb.sheetnames:
        if "OLAH" in sn.upper():
            continue
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        try:
            hdr = next(rows)
        except StopIteration:
            continue
        if not _is_pph_pull_sheet(hdr):
            continue
        cols = [str(c).strip() if c is not None else f"_c{i}"
                for i, c in enumerate(hdr)]
        data = [r for r in rows if any(c is not None for c in r)]
        if data:
            df = pd.DataFrame(data, columns=cols[:len(data[0])] if data else cols)
            df["_sheet"] = sn
            frames.append(df)
    wb.close()
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# ---------------------------------------------------------------- SIPO reader

SIPO_COLUMNS = ["Jenis Pajak", "Branch", "Masa Pajak", "Nama Vendor",
                "NPWP Vendor", "Jumlah Penghasilan", "Jumlah PPH",
                "Kode Objek Pajak", "Tarif", "Jenis Dok Reff", "No Dok Reff",
                "Tanggal Dok Reff"]


def _sipo_uker_stem(filename) -> str:
    """'152.xls' / '152(1).xls' / '152 (2).xls' -> '152'."""
    m = re.match(r"\s*(\d+)", os.path.basename(str(filename)))
    return m.group(1) if m else ""


def read_sipo(files) -> pd.DataFrame:
    """Consolidate the per-uker BRITAX exports (replaces Salsa's double
    Get Data). Accepts paths or uploaded file objects. The .xls files are
    legacy BIFF that xlrd flags as corrupt -> ignore_workbook_corruption.
    Re-downloaded duplicates ('152.xls' + '152(1).xls' with identical content)
    are dropped."""
    import xlrd
    frames = []
    seen = {}          # uker stem -> list of content fingerprints already kept
    for f in sorted(files, key=lambda x: len(getattr(x, "name", str(x)))):
        name = getattr(f, "name", str(f))
        if hasattr(f, "seek"):
            f.seek(0)
            book = xlrd.open_workbook(file_contents=f.read(),
                                      ignore_workbook_corruption=True)
        else:
            book = xlrd.open_workbook(f, ignore_workbook_corruption=True)
        sh = book.sheet_by_index(0)
        if sh.nrows < 2:
            continue
        hdr = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(1, sh.nrows)]
        stem = _sipo_uker_stem(name)
        fp = hash(str(rows))
        if stem and fp in seen.get(stem, []):
            continue                      # exact re-download of the same pull
        seen.setdefault(stem, []).append(fp)
        df = pd.DataFrame(rows, columns=hdr)
        df["_file"] = os.path.basename(name)
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def _masa_from_sipo(v):
    """'April 2026' -> (4, 2026)."""
    m = re.match(r"\s*([A-Za-z]+)\s+(\d{4})", str(v or ""))
    if not m:
        return None, None
    return BULAN_LONG.get(m.group(1).upper()), int(m.group(2))


# ---------------------------------------------------------------- builders

def _fmt_date(v):
    d = to_date(v)
    return d if d is not None else v


def _base_row(cfg: PphConfig, pasal: str) -> dict:
    return {
        "NPWP Pemotong": cfg.npwp_pemotong,
        "Masa Pajak": f"{cfg.masa:02d}",
        "Tahun Pajak": str(cfg.tahun),
        "Email": "",
        "Jenis PPh": PASAL[pasal]["jenis"],
        "Fasilitas Insentif": "9",
        "Nomor Setifikat Insentif": "-",
        "Tarif Fasilitas": "0",
        "Jenis Dokumen Referensi": "02",
        "Metode Pembayaran bagi Pemotong Instansi Pemerintah": "",
        "Nomor SP2D": "",
        "NPWP Penandatangan": cfg.npwp_penandatangan,
        "User Id": "",
        "Referensi 3": "", "Referensi 4": "", "Referensi 5": "",
    }


def build_template_sap(sap: pd.DataFrame, cfg: PphConfig, pasal: str) -> PphResult:
    """SAP pull rows -> PSIAP bukti-potong template for PPh 22 / 23 / 4A2."""
    tag = cfg.tag(pasal)
    out, exc, rec = [], [], []
    for i, r in sap.iterrows():
        nama = str(r.get("Nama") or "").strip()
        kop = str(r.get("KOP") or "").strip()
        # every pull row's tax counts toward the uker rekon (KOP or not — the
        # KOP-less deposito/reward rows still sit in the same GL)...
        cab = str(r.get("Kode Cabang Transaksi") or "").lstrip("0")
        try:
            uker = int(cab) if cab else None
        except ValueError:
            uker = None
        rec.append({"uker": uker, "pajak": to_num(r.get("Amt.in loc.cur."))})
        if not kop:
            # ...but rows without a KOP don't enter the template — surfaced so
            # Salsa can add back the few that do need a code.
            exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "KOP kosong",
                        "Keterangan": "Tidak masuk template — isi kode objek "
                                      "kalau memang perlu dilaporkan"})
            continue
        npwp = _npwp16(r.get("NIK/NPWP/TIN"))
        dok_inv = str(r.get("Dokumen Invoice") or "").strip()
        no_inv = str(r.get("Nomor Invoice") or "").strip()
        row = _base_row(cfg, pasal)

        nitku = norm_id(r.get("NITKU"))
        row["NITKU Pemotong (6 Digit Terakhir)"] = (
            nitku[-6:] if len(nitku) >= 6
            else _nitku_from_uker(str(r.get("Kode Cabang Transaksi") or "").lstrip("0")))
        if not row["NITKU Pemotong (6 Digit Terakhir)"]:
            exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "NITKU uker tidak ketemu",
                        "Keterangan": f"Kode cabang {r.get('Kode Cabang Transaksi')}"})

        if _zeros(npwp):
            row["NPWP Penerima Penghasilan"] = ""
            row["NITKU Penerima Penghasilan (22 Digit)"] = ""
            exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "NPWP penerima kosong/0000",
                        "Keterangan": "Konfirmasi ke uker lalu validasi di Coretax"})
        else:
            row["NPWP Penerima Penghasilan"] = npwp
            row["NITKU Penerima Penghasilan (22 Digit)"] = npwp + "000000"
            if npwp == cfg.npwp_pemotong:
                exc.append({"Baris": i + 2, "Vendor": nama,
                            "Jenis": "NPWP penerima = NPWP BRI",
                            "Keterangan": "SAP mengisi NPWP BRI sendiri — ganti NPWP vendor (cek Coretax)"})
            elif npwp == "1234567890123456" or len(set(npwp)) == 1:
                exc.append({"Baris": i + 2, "Vendor": nama,
                            "Jenis": "NPWP penerima mencurigakan",
                            "Keterangan": f"Nilai SAP '{npwp}' — validasi di Coretax"})
        row["Nama Penerima Penghasilan"] = nama

        row["Kode Objek Pajak"] = kop
        row["Penghasilan Bruto"] = round(to_num(r.get("DPP Amount Loc Currency WHT")))
        row["Nomor Dokumen Referensi"] = dok_inv if _zeros(no_inv) else no_inv
        row["Tanggal Dokumen Referensi"] = _fmt_date(r.get("Tanggal Invoice"))
        row["Tanggal Pemotongan"] = _fmt_date(r.get("Tanggal Pembayaran"))
        row["Referensi"] = f"{tag}_{dok_inv}"
        out.append(row)

    df = pd.DataFrame(out, columns=TEMPLATE_COLUMNS)
    return PphResult(
        template=df,
        exceptions=pd.DataFrame(exc, columns=["Baris", "Vendor", "Jenis", "Keterangan"]),
        recon_rows=pd.DataFrame(rec, columns=["uker", "pajak"]),
        stats={"pasal": pasal, "rows": len(df), "exceptions": len(exc)})


def build_template_sipo(sipo: pd.DataFrame, cfg: PphConfig) -> PphResult:
    """Consolidated BRITAX rows -> SIPOBRI template."""
    tag = cfg.tag("SIPOBRI")
    out, exc, rec = [], [], []
    seq = 0
    for i, r in sipo.iterrows():
        nama = str(r.get("Nama Vendor") or "").strip()
        npwp = _npwp16(r.get("NPWP Vendor"))
        row = _base_row(cfg, "SIPOBRI")

        masa, tahun = _masa_from_sipo(r.get("Masa Pajak"))
        if masa and (masa != cfg.masa or tahun != cfg.tahun):
            exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "Masa beda",
                        "Keterangan": f"File bilang {masa:02d}/{tahun}, run ini {cfg.masa:02d}/{cfg.tahun}"})

        # the pull is per uker: the FILE is named after the uker it was pulled
        # for (verified 180/180 vs the real template); the Branch column inside
        # is the transaction branch and can point elsewhere.
        uker_file = _sipo_uker_stem(r.get("_file")) or None
        row["NITKU Pemotong (6 Digit Terakhir)"] = (
            _nitku_from_uker(uker_file) or _nitku_from_uker(r.get("Branch")))
        if not row["NITKU Pemotong (6 Digit Terakhir)"]:
            exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "NITKU uker tidak ketemu",
                        "Keterangan": f"File {r.get('_file')} / branch {r.get('Branch')} tidak ada di master"})

        if _zeros(npwp):
            row["NPWP Penerima Penghasilan"] = ""
            row["NITKU Penerima Penghasilan (22 Digit)"] = ""
            exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "NPWP penerima kosong/0000",
                        "Keterangan": "Konfirmasi ke uker lalu validasi di Coretax"})
        else:
            row["NPWP Penerima Penghasilan"] = npwp
            row["NITKU Penerima Penghasilan (22 Digit)"] = npwp + "000000"
        row["Nama Penerima Penghasilan"] = nama

        kop = str(r.get("Kode Objek Pajak") or "").strip()
        if kop in ("", "-"):
            # older BRITAX pulls came without the KOP (Des 2025: all '-');
            # Salsa filled it by hand — surface instead of guessing.
            exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "Kode objek kosong",
                        "Keterangan": "BRITAX tidak mengisi KOP — isi kode objek pajak"})
        row["Kode Objek Pajak"] = kop
        row["Penghasilan Bruto"] = round(to_num(r.get("Jumlah Penghasilan")))
        row["Nomor Dokumen Referensi"] = str(r.get("No Dok Reff") or "").strip()
        row["Tanggal Dokumen Referensi"] = _fmt_date(r.get("Tanggal Dok Reff"))
        row["Tanggal Pemotongan"] = _fmt_date(r.get("Tanggal Dok Reff"))
        seq += 1
        row["Referensi"] = f"{tag}_{seq}"
        out.append(row)

        try:
            uker = int(uker_file) if uker_file else int(float(str(r.get("Branch")).strip()))
        except (TypeError, ValueError):
            uker = None
        rec.append({"uker": uker, "pajak": to_num(r.get("Jumlah PPH"))})

    df = pd.DataFrame(out, columns=TEMPLATE_COLUMNS)
    return PphResult(
        template=df,
        exceptions=pd.DataFrame(exc, columns=["Baris", "Vendor", "Jenis", "Keterangan"]),
        recon_rows=pd.DataFrame(rec, columns=["uker", "pajak"]),
        stats={"pasal": "SIPOBRI", "rows": len(df), "exceptions": len(exc)})


def _rate(v) -> float:
    """'1.5%' / '2%' / 0.015 -> 0.015."""
    s = str(v or "").strip().replace(",", ".")
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100
        except ValueError:
            return 0.0
    return to_num(v)


def build_data_olah_pph(sap: pd.DataFrame, cfg: PphConfig, pasal: str) -> pd.DataFrame:
    """Salsa's DATA OLAH view for one pasal: every SAP pull row + the derived
    uker/NITKU/referensi columns + the gap between the SAP setoran and the tax
    computed from the invoice (DPP x Rate) per baris."""
    tag = cfg.tag(pasal)
    rows = []
    for _, r in sap.iterrows():
        cab = str(r.get("Kode Cabang Transaksi") or "").lstrip("0")
        try:
            uker = int(cab) if cab else None
        except ValueError:
            uker = None
        uker_f = CABANG_INDUK.get(uker, uker) if uker is not None else None
        nitku = norm_id(r.get("NITKU"))
        dpp = to_num(r.get("DPP Amount Loc Currency WHT"))
        amt = to_num(r.get("Amt.in loc.cur."))
        rate = _rate(r.get("Rate Pajak"))
        hitung = round(dpp * rate)
        rows.append({
            "Kode Cabang Transaksi": r.get("Kode Cabang Transaksi"),
            "UKER": uker_f,
            "NAMA UKER": UKER_MASTER.get(uker_f, "") if uker_f is not None else "",
            "NITKU 6 DIGIT": (nitku[-6:] if len(nitku) >= 6
                              else UKER_NITKU.get(uker_f, "")),
            "REFERENSI": f"{tag}_{r.get('Dokumen Invoice') or ''}",
            "Dokumen Invoice": r.get("Dokumen Invoice"),
            "Dokumen Pembayaran": r.get("Dokumen Pembayaran"),
            "Dokumen Status": r.get("Dokumen Status"),
            "Masa Pajak": r.get("Masa Pajak"),
            "Tahun Pajak": r.get("Tahun Pajak"),
            "Tanggal Pembayaran": to_date(r.get("Tanggal Pembayaran")),
            "DPP (WHT)": dpp,
            "Rate Pajak": r.get("Rate Pajak"),
            "KOP": str(r.get("KOP") or "").strip(),
            "NIK/NPWP/TIN": r.get("NIK/NPWP/TIN"),
            "Nama": r.get("Nama"),
            "Nomor Invoice": r.get("Nomor Invoice"),
            "Tanggal Invoice": to_date(r.get("Tanggal Invoice")),
            "NITKU": r.get("NITKU"),
            "PPH HITUNG (DPP x Rate)": hitung,
            "SETORAN SAP (Amt)": amt,
            "SELISIH SETORAN vs HITUNG": round(amt + hitung),
            "CATATAN": ("KOP kosong — tidak masuk template"
                        if not str(r.get("KOP") or "").strip() else ""),
        })
    return pd.DataFrame(rows)


def read_dio_pph(path_or_buf) -> pd.DataFrame:
    """The DIO/SENDIK manual Excel (e.g. Pajak_PPH_23_Unifikasi_Manual_...xlsx):
    Jenis Pajak | Masa Pajak | Nama Vendor | NPWP Vendor | Jumlah Penghasilan |
    Jumlah PPH | Kanca | NPWP Uker | NITKU Uker | Kode Uker | Tarif | Kode Objek."""
    import openpyxl
    if hasattr(path_or_buf, "seek"):
        path_or_buf.seek(0)
    wb = openpyxl.load_workbook(path_or_buf, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        hi = next((i for i, r in enumerate(rows)
                   if r and any("NPWP Vendor" in str(c) for c in r if c)), None)
        if hi is None:
            continue
        cols = [str(c).strip() if c is not None else f"_c{i}"
                for i, c in enumerate(rows[hi])]
        data = [r for r in rows[hi + 1:] if r and any(c is not None for c in r)]
        wb.close()
        return pd.DataFrame(data, columns=cols[:len(data[0])] if data else cols)
    wb.close()
    return pd.DataFrame()


def build_template_dio(dio: pd.DataFrame, cfg: PphConfig, pasal: str = "23") -> PphResult:
    """DIO manual rows -> template. Everything pre-fills except the dokumen
    referensi fields — those live in the physical PDFs, so they stay blank and
    are flagged for Salsa. Jenis Dokumen Referensi = 02 like every other stream
    (the old 07 convention ended Jan 2026)."""
    tag = cfg.tag(pasal)
    out, exc, rec = [], [], []
    for i, r in dio.iterrows():
        nama = str(r.get("Nama Vendor") or "").strip()
        npwp = _npwp16(r.get("NPWP Vendor"))
        kode_uker = str(r.get("Kode Uker") or "").strip().lstrip("0")
        row = _base_row(cfg, pasal)

        row["NITKU Pemotong (6 Digit Terakhir)"] = _nitku_from_uker(kode_uker)
        if not row["NITKU Pemotong (6 Digit Terakhir)"]:
            nit = norm_id(r.get("NITKU Uker"))
            row["NITKU Pemotong (6 Digit Terakhir)"] = nit[-6:] if len(nit) >= 6 else ""
        if not row["NITKU Pemotong (6 Digit Terakhir)"]:
            exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "NITKU uker tidak ketemu",
                        "Keterangan": f"Kode Uker {r.get('Kode Uker')}"})

        if _zeros(npwp):
            row["NPWP Penerima Penghasilan"] = ""
            row["NITKU Penerima Penghasilan (22 Digit)"] = ""
            exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "NPWP penerima kosong/0000",
                        "Keterangan": "Konfirmasi ke uker lalu validasi di Coretax"})
        else:
            row["NPWP Penerima Penghasilan"] = npwp
            row["NITKU Penerima Penghasilan (22 Digit)"] = npwp + "000000"
        row["Nama Penerima Penghasilan"] = nama

        row["Kode Objek Pajak"] = str(r.get("Kode Objek Pajak") or "").strip()
        row["Penghasilan Bruto"] = round(to_num(r.get("Jumlah Penghasilan")))
        row["Nomor Dokumen Referensi"] = ""
        row["Tanggal Dokumen Referensi"] = ""
        row["Tanggal Pemotongan"] = ""
        exc.append({"Baris": i + 2, "Vendor": nama, "Jenis": "Dok referensi dari bukti fisik",
                    "Keterangan": "Isi nomor+tanggal dokumen & tanggal pemotongan dari PDF"})
        row["Referensi"] = tag
        out.append(row)

        try:
            uker = int(kode_uker) if kode_uker else None
        except ValueError:
            uker = None
        rec.append({"uker": uker, "pajak": -abs(to_num(r.get("Jumlah PPH")))})

    df = pd.DataFrame(out, columns=TEMPLATE_COLUMNS)
    return PphResult(
        template=df,
        exceptions=pd.DataFrame(exc, columns=["Baris", "Vendor", "Jenis", "Keterangan"]),
        recon_rows=pd.DataFrame(rec, columns=["uker", "pajak"]),
        stats={"pasal": f"{pasal} MANUAL", "rows": len(df), "exceptions": len(exc)})


# ---------------------------------------------------------------- rekon

# how each pasal's Utang column is labelled in the ETB PPh Unifikasi pivot
ETB_PPH_MARKER = {"22": "PASAL 22", "23": "PASAL 23", "4A2": "PASAL 4"}


def _uker_from_label(label, ro_name=None) -> int | None:
    """'00008 -- KC Baturaja' -> 8; 'KANWIL Palembang (Branch)' -> kanwil uker."""
    s = str(label or "").strip()
    m = re.match(r"\s*(\d+)\s*--", s)
    if m:
        return int(m.group(1))
    if "KANWIL" in s.upper() and ro_name:
        want = f"KANWIL {ro_name.strip().upper()}"
        for k, v in UKER_MASTER.items():
            if str(v).strip().upper() == want:
                try:
                    return int(float(str(k)))
                except (TypeError, ValueError):
                    continue
    return None


def read_etb_pph(path_or_buf, ro_name: str, pasal: str) -> dict:
    """uker -> Utang Pajak for one pasal, from the ETB PPh Unifikasi pivot
    (per-RO sheet, 'Utang Pajak - PPh Pasal NN - SAP' columns)."""
    import openpyxl
    if hasattr(path_or_buf, "seek"):
        path_or_buf.seek(0)
    wb = openpyxl.load_workbook(path_or_buf, read_only=True, data_only=True)
    ro = ro_name.strip().upper()
    marker = ETB_PPH_MARKER[pasal]

    def extract(ws):
        col = None
        vals = {}
        for r in ws.iter_rows(values_only=True):
            if col is None:
                if r and any(c and "PCA L2" in str(c) for c in r):
                    for cj, cc in enumerate(r):
                        if cc and marker in str(cc).upper():
                            col = cj
                            break
                    if col is None:
                        return {}
                continue
            label = r[0]
            if label is None or "GRAND TOTAL" in str(label).upper():
                continue
            uker = _uker_from_label(label, ro_name)
            if uker is None or col >= len(r) or r[col] is None:
                continue
            vals[uker] = vals.get(uker, 0.0) + to_num(r[col])
        return vals

    sheet = next((s for s in wb.sheetnames if s.strip().upper() == ro), None) \
        or next((s for s in wb.sheetnames if ro in s.upper()), None)
    out = extract(wb[sheet]) if sheet else {}
    if not out:
        # older files (e.g. per-RO ETB with a 'Data Olah' tab) have no RO-named
        # sheet — take the biggest uker->utang block anywhere in the workbook
        for sn in wb.sheetnames:
            cand = extract(wb[sn])
            if len(cand) > len(out):
                out = cand
    wb.close()
    return out


def build_rekon_pph(recon_rows: pd.DataFrame, utang: dict, ro_name: str = "") -> pd.DataFrame:
    """One pasal: PAJAK per uker (signed, negative like SAP Amt.in loc.cur.) vs
    the ETB Utang; SELISIH = UTANG - PAJAK, 0 when the ledger matches the pull.
    Mirrors Salsa's Sheet2. No branch reclass (all rows PEMBAYARAN)."""
    pajak = {}
    if recon_rows is not None:
        for _, r in recon_rows.iterrows():
            u = r["uker"]
            if u is None or pd.isna(u):
                continue
            u = CABANG_INDUK.get(int(u), int(u))
            pajak[u] = pajak.get(u, 0.0) + (r["pajak"] or 0.0)
    rows = []
    for u in sorted(set(utang) | set(pajak)):
        ut = utang.get(u)
        p = pajak.get(u)
        rows.append({
            "KODE UKER": u,
            "NAMA UKER": UKER_MASTER.get(str(u)) or UKER_MASTER.get(u) or "",
            "UTANG (ETB)": round(ut) if ut is not None else None,
            "PAJAK": round(p) if p is not None else None,
            "SELISIH": round((ut or 0) - (p or 0)),
        })
    return pd.DataFrame(rows, columns=["KODE UKER", "NAMA UKER", "UTANG (ETB)",
                                       "PAJAK", "SELISIH"])
