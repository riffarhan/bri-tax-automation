"""
Final validation across the full matrix: Palembang + Yogyakarta × April + May.

- April masa (WORK MAY 2026): full data exists -> validate template + recon
  against Salsa's real output.
- May masa (WORK JUNE 2026): only the SAP file exists (PPN work not done yet)
  -> dry-run (does the pipeline ingest + run end-to-end).
"""
import glob
import warnings

import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")
from engine import (Config, run, read_sap, read_coretax, read_etb,
                    build_cabang_index, build_rekon, build_doc_index,
                    coretax_seed_from_sap, normalize_coretax, reconcile,
                    sap_pull_sheets)

APR = "../WORK MAY 2026/"
JUN = "../WORK JUNE 2026/"


def g1(pattern):
    m = glob.glob(pattern)
    return m[0] if m else None


def template_check(sap, ct, expected, cfg):
    res = run(sap, ct, cfg, coretax_sheet="Faktur Masukan_1")
    exp = pd.read_excel(expected, sheet_name="FM - Import", dtype=object)
    exp = exp[exp["FM"] == "FM"].reset_index(drop=True)

    def kf(df):
        d = df.copy()
        for c in ["NOMOR_FAKTUR", "NPWP_WP", "ID_TKU_WP", "NPWP_PENJUAL",
                  "FIELD_TAMBAHAN_2", "FIELD_TAMBAHAN_1"]:
            d[c] = d[c].astype(str).str.strip()
        for c in ["KONFIRMASI", "MASA_PAJAK", "TAHUN_PAJAK",
                  "MASA_PENGKREDITAN", "TAHUN_PENGKREDITAN"]:
            d[c] = pd.to_numeric(d[c], errors="coerce").astype("Int64").astype(str)
        return d.set_index("NOMOR_FAKTUR")

    E, G = kf(exp), kf(res.fm_import)
    common = sorted(set(E.index) & set(G.index))
    cols = [c for c in E.columns if c in G.columns]
    tot = mm = crit_tot = crit_mm = 0
    for fk in common:
        for c in cols:
            tot += 1
            bad = str(E.loc[fk, c]) != str(G.loc[fk, c])
            mm += bad
            if c != "FIELD_TAMBAHAN_2":          # everything except the reference tag
                crit_tot += 1
                crit_mm += bad
    return (f"rows {len(G)}/{len(E)} | cells {100*(tot-mm)/tot:.1f}% | "
            f"tax-critical {100*(crit_tot-crit_mm)/crit_tot:.1f}% | "
            f"exceptions {res.stats['exceptions']}")


def recon_check(sap_path, ct, etb_path, cfg, etb_note=""):
    sap = read_sap(sap_path)
    coretax = read_coretax(ct, "Faktur Masukan_1")
    etb = read_etb(etb_path, ro_name=cfg.ro_name)
    cbf, cba, cbs = build_cabang_index(sap_path)
    rekon, flag = build_rekon(coretax, sap, etb, cfg, cbf, cba, cbs)
    wb = openpyxl.load_workbook(sap_path, data_only=True)
    if "REKON" not in wb.sheetnames:
        wb.close()
        return "no REKON sheet to compare"
    rk = wb["REKON"]
    hdr = [c.value for c in rk[1]]
    real = {}
    for i in range(2, rk.max_row + 1):
        u = rk.cell(i, 1).value
        if u in (None, ""):
            continue
        months = {hdr[j]: rk.cell(i, j + 1).value for j in range(1, 17)
                  if rk.cell(i, j + 1).value not in (None, 0)}
        real[int(u)] = {"months": months, "ETB": rk.cell(i, 19).value,
                        "SELISIH": rk.cell(i, 20).value}
    wb.close()
    gen = {int(r["KODE UKER"]): r for _, r in rekon.iterrows()}
    om = okm = oe = oke = osel = oksel = 0
    for u, info in real.items():
        gg = gen.get(u)
        if gg is None:
            continue
        for ml, v in info["months"].items():
            okm, om = (okm + 1, om) if (gg.get(ml) is not None and abs(gg[ml] - v) < 1) else (okm, om + 1)
        if info["ETB"] is not None:
            oke, oe = (oke + 1, oe) if (gg.get("ETB") is not None and abs(gg["ETB"] - info["ETB"]) < 1) else (oke, oe + 1)
        if info["SELISIH"] is not None:
            oksel, osel = (oksel + 1, osel) if (gg.get("SELISIH") is not None and abs(gg["SELISIH"] - info["SELISIH"]) < 1) else (oksel, osel + 1)
    return (f"ukers {len(gen)}/{len(real)} | months {okm}ok/{om}diff | "
            f"ETB {oke}ok/{oe}diff{etb_note} | SELISIH {oksel}ok/{osel}diff | flagged {len(flag)}")


def dry_run(sap_path, cfg):
    sap = read_sap(sap_path)
    inv = int((sap["dokumen_status"] == "INVOICE").sum())
    bf, ba = build_doc_index(sap_path)
    seed = coretax_seed_from_sap(sap, cfg)
    res = reconcile(normalize_coretax(seed), sap, cfg, bf, ba)
    return (f"sheets {sap_pull_sheets(sap_path)} | INVOICE {inv} | "
            f"doc-index {len(bf)} | template rows {res.stats['fm_import_rows']} (pipeline OK)")


print("=" * 78)
print("FINAL VALIDATION — PPN WAPU (Alfa)")
print("=" * 78)

# ---- APRIL (full data) ----
for ro, sap_glob, ct_glob, tpl_glob, etbnote in [
    ("PALEMBANG",
     APR + "PPN PALEMBANG/21 APR*SAP RO PALEMBANG*.xlsx",
     APR + "PPN PALEMBANG/EKSPOR PSIAP PPN WAPU PALEMBANG*.xlsx",
     APR + "PPN PALEMBANG/Template Impor*PALEMBANG*- PSIAP.xlsx", ""),
    ("YOGYAKARTA",
     APR + "PPN YOGYAKARTA/21 APR*SAP RO YOGYAKAR*.xlsx",
     APR + "PPN YOGYAKARTA/EKSPOR PSIAP PPN WAPU RO YOGYAKARTA*.xlsx",
     APR + "PPN YOGYAKARTA/Template Impor*YOGYAKARTA*- PSIAP.xlsx",
     " (real REKON used a different ETB snapshot not in set)"),
]:
    cfg = Config(ro_name=ro, masa=4, tahun=2026)
    sap, ct, tpl = g1(sap_glob), g1(ct_glob), g1(tpl_glob)
    etb = g1(APR + "ETB PPN WAPU*.xlsx")
    print(f"\n### APRIL · {ro}")
    print("  TEMPLATE:", template_check(sap, ct, tpl, cfg))
    print("  REKON   :", recon_check(sap, ct, etb, cfg, etbnote))

# ---- MAY (SAP only -> dry-run) ----
for ro, sap_glob in [
    ("PALEMBANG", JUN + "PPN PALEMBANG/21 may*SAP RO PALEMBANG*.xlsx"),
    ("YOGYAKARTA", JUN + "PPN YOGYAKARTA/21 may*SAP RO YOGYAKARTA*.xlsx"),
]:
    cfg = Config(ro_name=ro, masa=5, tahun=2026)
    sap = g1(sap_glob)
    print(f"\n### MAY · {ro}  (PPN work not done yet — dry-run only)")
    print("  DRY-RUN :", dry_run(sap, cfg) if sap else "SAP file not found")

print("\n" + "=" * 78)
