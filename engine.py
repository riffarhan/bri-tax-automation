"""
PPN WAPU PSIAP — transformation & reconciliation engine.

Pilot scope: RO Palembang. Produces the PSIAP `FM-Import` template plus an
exception report, from two inputs:

  1. SAP Pajak Masukan extract  (the "...SAP RO..." workbook, sheet "Sheet1")
  2. Coretax Pajak Masukan list (NPWP penjual, faktur, masa, tahun, dpp, ppn,
     status, konfirmasi, valid) — supplied today as an upload; later fetched
     automatically by the Coretax connector (RPA).

Design notes (validated against April 2026 Palembang, see validate_april.py):
  * FM-Import is a 1:1 reshape of the Coretax faktur set. masa/tahun/npwp_penjual
    all come from Coretax — SAP does NOT carry the faktur masa.
  * SAP is used for reconciliation: the FT2 doc-number tag (Dokumen Invoice,
    joined by faktur) and to surface discrepancies (DPP diff, booked-not-reported).
  * Nothing is silently dropped: ambiguous rows go to the exception report.
"""
from __future__ import annotations
import csv
import datetime as dt
import os
import re
from dataclasses import dataclass, field

import pandas as pd

# ----------------------------------------------------------------------------- config / constants
BULAN_ID = {1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL", 5: "MEI",
            6: "JUNI", 7: "JULI", 8: "AGUSTUS", 9: "SEPTEMBER", 10: "OKTOBER",
            11: "NOVEMBER", 12: "DESEMBER"}

FM_IMPORT_COLUMNS = ["FM", "NPWP_WP", "ID_TKU_WP", "NOMOR_FAKTUR", "KONFIRMASI",
                     "MASA_PAJAK", "TAHUN_PAJAK", "MASA_PENGKREDITAN",
                     "TAHUN_PENGKREDITAN", "NPWP_PENJUAL",
                     "FIELD_TAMBAHAN_1", "FIELD_TAMBAHAN_2"]
# the "full" review export = the upload columns + these context/recon columns
FULL_EXTRA_COLUMNS = ["Nama Vendor", "Document Number SAP", "Kode Uker", "Nama Uker",
                      "Status Coretax", "DPP", "Tarif", "Jumlah Pajak di Faktur",
                      "Jumlah Pajak di SAP", "Selisih", "Tanggal Faktur"]
FM_IMPORT_FULL_COLUMNS = FM_IMPORT_COLUMNS + FULL_EXTRA_COLUMNS

# Coretax crediting status -> PSIAP KONFIRMASI code, confirmed by the Petunjuk
# Isian: 1 = Dikreditkan (credited), 2 = Tidak Dikreditkan (uncredited). The grid
# keeps the raw Coretax value ('uncredited'/'credited') as in the Faktur Masukan
# sheet; the upload template gets the code (uncredited -> 2).
KONFIRMASI_CODE = {"uncredited": 2, "credited": 1}


def _load_uker_master():
    """kode_uker -> nama_uker, from the bundled slim master (names only)."""
    path = os.path.join(os.path.dirname(__file__), "data", "uker_master.csv")
    out = {}
    try:
        with open(path, newline="") as f:
            for row in csv.DictReader(f):
                out[int(row["kode_uker"])] = row["nama_uker"]
    except (OSError, ValueError, KeyError):
        pass
    return out


UKER_MASTER = _load_uker_master()   # exact uker names (master); ETB is the fallback


@dataclass
class Config:
    ro_name: str = "PALEMBANG"          # appears in the FT2 label
    masa: int = 4                        # filing masa (month) — drives the label
    tahun: int = 2026
    npwp_wp: str = "0010016087093000"    # BRI HO
    id_tku_wp: str = "000000"
    branch_tag: str = "HO"               # FIELD_TAMBAHAN_1
    label_prefix: str = "PPNWAPUSAPRO"   # full label = prefix + RO + MONTH + YEAR

    @property
    def label(self) -> str:
        return f"{self.label_prefix}{self.ro_name}{BULAN_ID[self.masa]}{self.tahun}"


# ----------------------------------------------------------------------------- normalisation helpers
def norm_id(v) -> str:
    """Normalise an NPWP/NITKU/TIN-like value to a clean digit string."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if isinstance(v, float):                       # e.g. 1.0016087093e+19
        v = f"{v:.0f}"
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def norm_faktur(v) -> str:
    s = norm_id(v)
    return re.sub(r"\D", "", s)                     # keep digits only


def faktur_key(faktur: str, n: int = 6) -> str:
    """Dedup key — Coretax flags only 'similar' fakturs, so Salsa matches on the
    last few digits. Default 6 (between her 5–7)."""
    f = norm_faktur(faktur)
    return f[-n:] if len(f) >= n else f


def to_month(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.month
    if isinstance(v, str):
        m = re.match(r"\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", v)
        if m:                                       # dd/mm/yyyy
            return int(m.group(2))
    return None


def to_num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def to_date(v):
    """Parse a date value or 'dd/mm/yyyy' string to a date; else None."""
    if isinstance(v, (dt.datetime, dt.date)):
        return v
    if isinstance(v, str):
        m = re.match(r"\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", v)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            y += 2000 if y < 100 else 0
            try:
                return dt.date(y, mo, d)
            except ValueError:
                return None
    return None


# ----------------------------------------------------------------------------- readers
SAP_RENAME = {
    "Kode Cabang Transaksi": "kode_cabang", "Dokumen Invoice": "dokumen_invoice",
    "Dokumen Pembayaran": "dokumen_pembayaran", "Dokumen Status": "dokumen_status",
    "Masa Pajak": "masa_sap", "Tahun Pajak": "tahun_sap",
    "DPP Amount Loc Currency VAT": "dpp_sap", "Amt.in loc.cur.": "amt_loc",
    "NIK/NPWP/TIN": "npwp_sap",
    "Nama": "nama_sap", "Tanggal Faktur": "tgl_faktur_sap", "Nomor FP": "nomor_faktur",
}

# Canonical Coretax columns the engine consumes, and the aliases it accepts
# (the April validation source is the "Faktur Masukan" sheet).
CORETAX_ALIASES = {
    "npwp_penjual": ["NPWP_PENJUAL", "npwp_penjual"],
    "nama_penjual": ["NAMA_PENJUAL", "nama_penjual", "NAMA"],
    "nomor_faktur": ["NOMOR_FAKTUR", "nomor_faktur"],
    "tanggal_faktur": ["TANGGAl_FAKTUR", "TANGGAL_FAKTUR", "tanggal_faktur"],
    "masa": ["MASA", "masa"],
    "tahun": ["TAHUN", "tahun"],
    "dpp": ["DPP", "dpp"],
    "ppn": ["PPN", "ppn"],
    "status": ["STATUS", "status"],
    "konfirmasi": ["KONFIRMASI", "konfirmasi"],
    "valid": ["VALID", "valid"],
    "branch": ["BRANCH", "branch", "FIELD_TAMBAHAN_1"],
}


def _read_sap_sheet(path, sheet) -> pd.DataFrame:
    """Read one SAP sheet, auto-detecting the header row (Sheet1 = row 1,
    BULAN SBLMNYA / BERIKUTNYA = row 2)."""
    probe = pd.read_excel(path, sheet_name=sheet, header=None, nrows=3, dtype=object)
    hdr_row = 0
    for i in range(min(3, len(probe))):
        if probe.iloc[i].astype(str).str.strip().eq("Nomor FP").any():
            hdr_row = i
            break
    df = pd.read_excel(path, sheet_name=sheet, header=hdr_row, dtype=object)
    df = df.rename(columns={k: v for k, v in SAP_RENAME.items() if k in df.columns})
    keep = [c for c in SAP_RENAME.values() if c in df.columns]
    df = df[keep].copy()
    df["_sheet"] = sheet
    for c in ("nomor_faktur",):
        if c in df: df[c] = df[c].map(norm_faktur)
    for c in ("npwp_sap", "dokumen_invoice"):
        if c in df: df[c] = df[c].map(norm_id)
    for c in ("dpp_sap", "amt_loc"):
        if c in df: df[c] = df[c].map(to_num)
    return df


_SAP_MARKERS = {"Nomor FP", "Dokumen Invoice"}          # columns the doc-index needs


def _is_doc_source_sheet(path, sheet) -> bool:
    """True for any sheet carrying faktur→doc data — detected by having both
    'Nomor FP' and 'Dokumen Invoice' columns. Catches Sheet1, partial-download
    tabs (19-20, 15-19, 21-21, …) and the adjacent-month worked sheets
    (BULAN SBLMNYA/BERIKUTNYA), which hold the doc numbers for fakturs booked in
    a neighbouring month. REKON / DUPLIKAT lack these columns and drop out.
    Name-agnostic, so it survives month-to-month tab renames."""
    try:
        probe = pd.read_excel(path, sheet_name=sheet, header=None, nrows=3, dtype=object)
    except Exception:
        return False
    for i in range(min(3, len(probe))):
        vals = {str(x).strip() for x in probe.iloc[i].tolist()}
        if _SAP_MARKERS <= vals:
            return True
    return False


def sap_pull_sheets(path) -> list:
    import openpyxl
    names = openpyxl.load_workbook(path, read_only=True).sheetnames
    # 'DATA OLAH' is the current-month OUTPUT — exclude it so we don't reuse
    # Salsa's own answers (it's empty on a fresh month anyway). BULAN
    # SBLMNYA/BERIKUTNYA are kept: they hold real adjacent-month doc numbers.
    return [s for s in names
            if "OLAH" not in s.upper() and _is_doc_source_sheet(path, s)]


def read_sap(path, sheet="Sheet1") -> pd.DataFrame:
    """Current-masa SAP rows, used for reconciliation (DPP diff, booked-not-reported)."""
    import openpyxl
    available = openpyxl.load_workbook(path, read_only=True).sheetnames
    if sheet not in available:                       # fall back to the first SAP pull tab
        pulls = sap_pull_sheets(path)
        sheet = pulls[0] if pulls else available[0]
    return _read_sap_sheet(path, sheet)


def build_doc_index(path):
    """Faktur/doc lookup spanning every SAP pull tab (current + adjacent-month +
    partial downloads), since a faktur reported this masa may have been booked
    in an adjacent month."""
    frames = [_read_sap_sheet(path, s) for s in sap_pull_sheets(path)]
    if not frames:
        return {}, {}
    allrows = pd.concat(frames, ignore_index=True)
    allrows = allrows[allrows.get("dokumen_invoice", "").astype(str) != ""]
    by_faktur = {r["nomor_faktur"]: r["dokumen_invoice"]
                 for _, r in allrows.iterrows() if r.get("nomor_faktur")}
    # NPWP + DPP (rounded) -> doc, only when unambiguous
    by_amt = {}
    for _, r in allrows.iterrows():
        k = (norm_id(r.get("npwp_sap")), round(to_num(r.get("dpp_sap"))))
        if not k[0] or k[1] == 0:
            continue
        by_amt.setdefault(k, set()).add(r["dokumen_invoice"])
    by_amt = {k: next(iter(v)) for k, v in by_amt.items() if len(v) == 1}
    return by_faktur, by_amt


def resolve_doc(faktur, npwp, dpp, by_faktur, by_amt):
    """Three-tier: exact faktur -> NPWP+amount -> none. Returns (doc, how)."""
    if faktur in by_faktur:
        return by_faktur[faktur], "faktur"
    k = (norm_id(npwp), round(to_num(dpp)))
    if k in by_amt:
        return by_amt[k], "npwp+nominal"
    return "", "tidak ketemu"


def normalize_coretax(raw: pd.DataFrame) -> pd.DataFrame:
    """Canonicalise a Coretax table (from a file or a pasted/edited grid)."""
    if "FM" in raw.columns:                          # 'Faktur Masukan' sheet marker
        raw = raw[raw["FM"] == "FM"]
    out = {}
    for canon, aliases in CORETAX_ALIASES.items():
        col = next((a for a in aliases if a in raw.columns), None)
        out[canon] = raw[col].values if col is not None else None
    df = pd.DataFrame(out)
    df["nomor_faktur"] = df["nomor_faktur"].map(norm_faktur)
    df["npwp_penjual"] = df["npwp_penjual"].map(norm_id)
    df["dpp"] = df["dpp"].map(to_num)
    df["ppn"] = df["ppn"].map(to_num)
    df = df[df["nomor_faktur"] != ""].reset_index(drop=True)
    return df


def read_coretax(path, sheet=0) -> pd.DataFrame:
    return normalize_coretax(pd.read_excel(path, sheet_name=sheet, dtype=object))


def coretax_seed_from_sap(sap: pd.DataFrame, cfg: "Config") -> pd.DataFrame:
    """Pre-fill an editable grid from SAP so Salsa mostly only types the masa
    (and fixes status) she reads off Coretax — not the whole row."""
    inv = sap[sap["dokumen_status"] == "INVOICE"]
    return pd.DataFrame({
        "nomor_faktur": inv["nomor_faktur"].astype(str).values,
        "npwp_penjual": inv["npwp_sap"].astype(str).values,
        "nama_penjual": inv["nama_sap"].astype(str).values,
        "masa": [None] * len(inv),                   # <- from Coretax (key field)
        "tahun": [cfg.tahun] * len(inv),
        "dpp": inv["dpp_sap"].values,
        "ppn": [None] * len(inv),                    # <- from Coretax
        "status": ["approved"] * len(inv),
        "konfirmasi": ["uncredited"] * len(inv),
    })


# ----------------------------------------------------------------------------- core
@dataclass
class Result:
    fm_import: pd.DataFrame
    exceptions: pd.DataFrame
    stats: dict = field(default_factory=dict)
    fm_import_full: pd.DataFrame = None


def reconcile(coretax: pd.DataFrame, sap: pd.DataFrame, cfg: Config,
              by_faktur=None, by_amt=None,
              cab_by_faktur=None, cab_by_amt=None, cab_by_suffix=None,
              uker_names=None) -> Result:
    by_faktur = by_faktur or {}
    by_amt = by_amt or {}
    cab_by_faktur = cab_by_faktur or {}
    cab_by_amt = cab_by_amt or {}
    cab_by_suffix = cab_by_suffix or {}
    uker_names = uker_names or {}
    etb_ukers = set(uker_names)
    reclass = build_reclass_map(sap, etb_ukers) if etb_ukers else {}
    sap_inv = sap[sap["dokumen_status"] == "INVOICE"].copy()
    sap_by_faktur = {r["nomor_faktur"]: r for _, r in sap_inv.iterrows()
                     if r["nomor_faktur"]}

    rows, full, exc = [], [], []

    # duplicate detection (last-6-digit key) within the Coretax set
    dup_keys = (coretax.assign(_k=coretax["nomor_faktur"].map(faktur_key))
                .groupby("_k")["nomor_faktur"].transform("nunique"))
    coretax = coretax.assign(_dupgroup=coretax["nomor_faktur"].map(faktur_key),
                             _dupcount=dup_keys)

    for _, c in coretax.iterrows():
        fk = c["nomor_faktur"]
        s = sap_by_faktur.get(fk)
        dok, how = resolve_doc(fk, c["npwp_penjual"], c["dpp"], by_faktur, by_amt)

        # --- build the FM-Import row (every Coretax faktur is reportable) ---
        base = {
            "FM": "FM",
            "NPWP_WP": cfg.npwp_wp,
            "ID_TKU_WP": cfg.id_tku_wp,
            "NOMOR_FAKTUR": fk,
            "KONFIRMASI": KONFIRMASI_CODE.get(str(c["konfirmasi"]).strip().lower(), 2),
            "MASA_PAJAK": int(c["masa"]) if pd.notna(c["masa"]) else None,
            "TAHUN_PAJAK": int(c["tahun"]) if pd.notna(c["tahun"]) else None,
            "MASA_PENGKREDITAN": int(c["masa"]) if pd.notna(c["masa"]) else None,
            "TAHUN_PENGKREDITAN": int(c["tahun"]) if pd.notna(c["tahun"]) else None,
            "NPWP_PENJUAL": c["npwp_penjual"],
            "FIELD_TAMBAHAN_1": (norm_id(c["branch"]) or cfg.branch_tag),
            "FIELD_TAMBAHAN_2": f"{cfg.label}_{dok}" if dok else f"{cfg.label}_",
        }
        rows.append(base)

        # --- full review row (upload columns + context/recon columns) ---
        info = resolve_sap_info(fk, c["npwp_penjual"], c.get("dpp"),
                                cab_by_faktur, cab_by_amt, cab_by_suffix)
        uker = derive_uker(info.get("cabang"))
        if uker is not None and uker not in etb_ukers:
            uker = reclass.get(uker, uker)
        ppn_fak = to_num(c.get("ppn"))
        if info:
            dpp_v, sap_ppn = to_num(info.get("dpp")), to_num(info.get("amt"))
        else:
            dpp_v, sap_ppn = to_num(c.get("dpp")), None
        full.append({**base,
            "Nama Vendor": c.get("nama_penjual"),
            "Document Number SAP": dok,
            "Kode Uker": uker,
            "Nama Uker": (UKER_MASTER.get(uker) or uker_names.get(uker)) if uker is not None else None,
            "Status Coretax": str(c.get("status") or "").upper(),
            "DPP": round(dpp_v) if dpp_v else None,
            "Tarif": round(ppn_fak / dpp_v, 4) if dpp_v else None,
            "Jumlah Pajak di Faktur": round(ppn_fak) if ppn_fak else None,
            "Jumlah Pajak di SAP": round(sap_ppn) if sap_ppn else None,
            "Selisih": round(ppn_fak + sap_ppn) if sap_ppn is not None else None,
            "Tanggal Faktur": to_date(c.get("tanggal_faktur")),
        })

        # --- exception checks (flag for review, never block) ---
        blank_faktur = (not fk) or set(fk) <= {"0"}
        if blank_faktur:
            exc.append(_ex(fk, c, "Nomor faktur kosong",
                           "Nomor faktur kosong / semua nol di sumber — isi nomor e-faktur yang benar dari Coretax."))
        if how == "tidak ketemu":
            exc.append(_ex(fk, c, "Doc number tidak ketemu",
                           "Tidak ketemu di SAP via faktur maupun NPWP+nominal — doc number perlu diisi / dicek manual."))
        elif how == "npwp+nominal":
            exc.append(_ex(fk, c, "Doc via NPWP+nominal",
                           f"Faktur tidak persis cocok di SAP (kemungkinan typo); doc {dok} dicocokkan via NPWP+nominal — mohon verifikasi."))
        if s is not None:
            diff = to_num(c["dpp"]) - to_num(s["dpp_sap"])
            if abs(diff) >= 1:
                exc.append(_ex(fk, c, "DPP beda",
                               f"DPP Coretax {to_num(c['dpp']):,.0f} vs SAP {to_num(s['dpp_sap']):,.0f} (selisih {diff:,.0f})"))
        if str(c["status"]).strip().lower() != "approved":
            exc.append(_ex(fk, c, "Belum approved",
                           f"Status Coretax = {c['status']} — faktur belum bisa dikreditkan"))
        if c["_dupcount"] > 1 and not blank_faktur:
            exc.append(_ex(fk, c, "Dugaan duplikat",
                           f"Nomor faktur mirip (6 digit akhir = {c['_dupgroup']}) dengan faktur lain"))

    # SAP fakturs that never appear in Coretax = booked but not reported
    coretax_keys = set(coretax["nomor_faktur"])
    for fk, s in sap_by_faktur.items():
        if fk not in coretax_keys:
            exc.append({"Nomor Faktur": fk, "NPWP Penjual": s["npwp_sap"],
                        "Nama": s.get("nama_sap"), "Masa": s.get("masa_sap"),
                        "Jenis": "Tidak ada di Coretax",
                        "Keterangan": "Ada di SAP (INVOICE) tapi tidak ketemu di Coretax — cek apakah perlu dilaporkan / beda masa"})

    fm_import = pd.DataFrame(rows, columns=FM_IMPORT_COLUMNS)
    fm_import_full = pd.DataFrame(full, columns=FM_IMPORT_FULL_COLUMNS)
    exceptions = pd.DataFrame(exc, columns=["Nomor Faktur", "NPWP Penjual", "Nama",
                                            "Masa", "Jenis", "Keterangan"])
    stats = {
        "coretax_fakturs": len(coretax),
        "sap_invoice_rows": len(sap_inv),
        "fm_import_rows": len(fm_import),
        "doc_filled": sum(1 for r in rows if r["FIELD_TAMBAHAN_2"] != f"{cfg.label}_"),
        "exceptions": len(exceptions),
    }
    return Result(fm_import, exceptions, stats, fm_import_full)


def _ex(fk, c, jenis, ket):
    return {"Nomor Faktur": fk, "NPWP Penjual": c["npwp_penjual"],
            "Nama": c.get("nama_penjual"), "Masa": c.get("masa"),
            "Jenis": jenis, "Keterangan": ket}


def run(sap_path, coretax_path, cfg: Config,
        sap_sheet="Sheet1", coretax_sheet=0, etb_path=None) -> Result:
    sap = read_sap(sap_path, sap_sheet)
    by_faktur, by_amt = build_doc_index(sap_path)
    cbf, cba, cbs = build_cabang_index(sap_path)
    coretax = read_coretax(coretax_path, coretax_sheet)
    uker_names = read_uker_names(etb_path, ro_name=cfg.ro_name) if etb_path else {}
    return reconcile(coretax, sap, cfg, by_faktur, by_amt, cbf, cba, cbs, uker_names)


# ----------------------------------------------------------------------------- recon (REKON)
# Validated against the real April Palembang REKON: month value = sum of the
# Coretax faktur PPN per uker per masa; ETB = per-uker ledger balance; and
# SELISIH = ETB + TOTAL.
REKON_MONTHS = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", "JULI",
                "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
ETB_SHEET_BY_RO = {"PALEMBANG": "PLG", "YOGYAKARTA": "YOG"}


def _extract_uker_balance(sh) -> dict:
    """Find the (uker, balance) helper block in an ETB sheet by scanning for the
    adjacent column pair that yields the most (small-int uker, number) rows —
    robust to the columns shifting between files/ROs."""
    rows = list(sh.iter_rows(values_only=True))
    ncol = max((len(r) for r in rows), default=0)
    best = {}
    for c in range(ncol - 1):
        out = {}
        for r in rows:
            if c + 1 >= len(r):
                continue
            u, bal = r[c], r[c + 1]
            if (isinstance(u, (int, float)) and not isinstance(u, bool)
                    and 1 <= u <= 99999 and float(u) == int(u)
                    and isinstance(bal, (int, float)) and not isinstance(bal, bool)):
                out[int(u)] = float(bal)
        if len(out) > len(best):
            best = out
    return best


def read_etb(path, sheet=None, ro_name=None) -> dict:
    """uker(int) -> ledger balance, from the ETB workbook. Picks the per-RO sheet
    (PLG/YOG) when present; otherwise scans every sheet for the uker→balance block."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target = sheet or (ETB_SHEET_BY_RO.get(ro_name.strip().upper()) if ro_name else None)
    if target and target in wb.sheetnames:
        out = _extract_uker_balance(wb[target])
    else:
        out = {}
        for sn in wb.sheetnames:
            cand = _extract_uker_balance(wb[sn])
            if len(cand) > len(out):
                out = cand
    wb.close()
    return out


def read_uker_names(path, sheet=None, ro_name=None) -> dict:
    """uker(int) -> uker name, parsed from ETB labels like '00020 -- KC Jambi'
    (the leading digits are the uker, the part after ' -- ' is the name)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target = sheet or (ETB_SHEET_BY_RO.get(ro_name.strip().upper()) if ro_name else None)
    sheets = [target] if (target and target in wb.sheetnames) else wb.sheetnames
    out = {}
    for sn in sheets:
        for row in wb[sn].iter_rows(values_only=True):
            for i, v in enumerate(row):
                if isinstance(v, str) and " -- " in v:                # '00342 -- KC ...'
                    prefix, name = v.split(" -- ", 1)
                    pn = "".join(ch for ch in prefix if ch.isdigit())
                    if pn and name.strip():
                        out.setdefault(int(pn), name.strip())
                elif (isinstance(v, (int, float)) and not isinstance(v, bool)
                      and 1 <= v <= 99999 and float(v) == int(v)):     # code + plain label
                    for j in (i - 1, i + 1):
                        if (0 <= j < len(row) and isinstance(row[j], str)
                                and row[j].strip() and not row[j].strip()[:6].strip().isdigit()):
                            out.setdefault(int(v), row[j].strip())
                            break
        if out:
            break
    wb.close()
    return out


def derive_uker(kode_cabang):
    """Parent uker from the SAP branch code (strip leading zeros). Returns None
    for branch/non-numeric codes that need a reclass to their parent."""
    s = norm_id(kode_cabang)
    return int(s) if s.isdigit() else None


def build_reclass_map(sap: pd.DataFrame, etb_ukers: set) -> dict:
    """branch uker -> parent uker, from the SAP RECLASS pairs. A branch (a uker
    not among the known ETB/parent ukers) entered with +X is matched by equal
    |amount| to a parent entered with -X. This is the manual 'match by nominal'."""
    from collections import defaultdict
    rec = sap[sap["dokumen_status"] == "RECLASS"]
    by_amt = defaultdict(list)
    for _, r in rec.iterrows():
        u = derive_uker(r.get("kode_cabang"))
        amt = to_num(r.get("amt_loc"))
        if u is not None and amt:
            by_amt[round(abs(amt))].append(u)
    mapping = {}
    for _, ukers in by_amt.items():
        branches = [u for u in ukers if u not in etb_ukers]
        parents = [u for u in ukers if u in etb_ukers]
        if len(branches) == 1 and len(parents) == 1:
            mapping[branches[0]] = parents[0]
    return mapping


def build_cabang_index(path):
    """faktur->kode_cabang and (npwp,amt)->kode_cabang across ALL SAP pull sheets
    (so a faktur booked in an adjacent month still resolves its uker)."""
    from collections import defaultdict
    frames = [_read_sap_sheet(path, s) for s in sap_pull_sheets(path)]
    if not frames:
        return {}, {}
    allrows = pd.concat(frames, ignore_index=True)
    allrows = allrows[allrows.get("dokumen_status", "") == "INVOICE"]
    by_faktur = {}
    by_amt, by_suffix = defaultdict(list), defaultdict(list)
    for _, r in allrows.iterrows():
        info = {"cabang": r.get("kode_cabang"),
                "dpp": to_num(r.get("dpp_sap")), "amt": to_num(r.get("amt_loc"))}
        fk = r.get("nomor_faktur")
        if fk:
            by_faktur.setdefault(fk, info)
            by_suffix[faktur_key(fk, 7)].append(info)
        k = (norm_id(r.get("npwp_sap")), round(to_num(r.get("dpp_sap"))))
        if k[0] and k[1]:
            by_amt[k].append(info)
    # keep a key only if it maps to ONE uker (cabang); store the first row's info
    by_amt = {k: v[0] for k, v in by_amt.items() if len({i["cabang"] for i in v}) == 1}
    by_suffix = {k: v[0] for k, v in by_suffix.items() if len({i["cabang"] for i in v}) == 1}
    return by_faktur, by_amt, by_suffix


def resolve_sap_info(faktur, npwp, dpp, by_faktur, by_amt, by_suffix) -> dict:
    """Resolve a Coretax faktur to its SAP info {cabang, dpp, amt}:
    exact faktur -> last-7 suffix -> NPWP+amount. {} if none."""
    return (by_faktur.get(faktur)
            or by_suffix.get(faktur_key(faktur, 7))
            or by_amt.get((norm_id(npwp), round(to_num(dpp))))
            or {})


def build_rekon(coretax: pd.DataFrame, sap: pd.DataFrame, etb: dict, cfg: Config,
                cab_by_faktur=None, cab_by_amt=None, cab_by_suffix=None):
    """Pivot reportable PPN by uker × masa, fold branches into their parent (via
    the RECLASS map), join the ETB balance, compute SELISIH. Lists every ETB uker.
    Fakturs whose uker can't be resolved are flagged. Returns (rekon_df, flagged_df).
    Pass the cab_* maps from build_cabang_index() to resolve uker across sheets."""
    from collections import defaultdict
    etb_ukers = set(etb)
    reclass = build_reclass_map(sap, etb_ukers)

    if cab_by_faktur is None:
        inv = sap[sap["dokumen_status"] == "INVOICE"]
        cab_by_faktur = {r["nomor_faktur"]: {"cabang": r.get("kode_cabang"),
                                             "dpp": to_num(r.get("dpp_sap")),
                                             "amt": to_num(r.get("amt_loc"))}
                         for _, r in inv.iterrows() if r["nomor_faktur"]}
    cab_by_amt = cab_by_amt or {}
    cab_by_suffix = cab_by_suffix or {}

    pivot = defaultdict(float)
    flagged = []
    for _, c in coretax.iterrows():
        fk = c["nomor_faktur"]
        ppn = to_num(c.get("ppn"))
        masa = c.get("masa")
        info = resolve_sap_info(fk, c["npwp_penjual"], c.get("dpp"),
                                cab_by_faktur, cab_by_amt, cab_by_suffix)
        cab = info.get("cabang")
        uker = derive_uker(cab)
        if uker is not None and uker not in etb_ukers:
            uker = reclass.get(uker, uker)               # fold branch into parent
        if uker is None or uker not in etb_ukers:
            flagged.append({"Nomor Faktur": fk, "Kode Cabang": cab, "PPN": round(ppn),
                            "Keterangan": "Uker tidak terpetakan — cabang perlu reclass ke induk, atau faktur tidak ketemu di SAP"})
            continue
        if pd.notna(masa):
            pivot[(uker, int(masa))] += ppn

    rows = []
    for u in sorted(etb_ukers):
        row, total = {"KODE UKER": u}, 0.0
        for mi, mlabel in enumerate(REKON_MONTHS, 1):
            v = pivot.get((u, mi), 0.0)
            row[mlabel] = round(v) if v else None
            total += v
        e = etb.get(u)
        row["TOTAL"] = round(total)
        row["ETB"] = round(e) if e is not None else None
        row["SELISIH"] = round((e or 0) + total) if e is not None else None
        rows.append(row)
    cols = ["KODE UKER"] + REKON_MONTHS + ["TOTAL", "ETB", "SELISIH"]
    rekon_df = pd.DataFrame(rows, columns=cols).dropna(axis=1, how="all")
    flagged_df = pd.DataFrame(flagged, columns=["Nomor Faktur", "Kode Cabang", "PPN", "Keterangan"])
    return rekon_df, flagged_df
