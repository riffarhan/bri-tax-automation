"""
Stress-test Alfa across every historical masa with complete PPN data
(Des 2025 ... Apr 2026, both ROs). Validates template + recon + full version
against the real GRUP/PSIAP files, and surfaces any edge case / error.
"""
import glob
import os
import traceback
import warnings

import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")
from engine import (Config, read_sap, read_coretax, read_etb, build_cabang_index,
                    build_rekon, run, REKON_MONTHS)

ROOT = ".."
# (work folder, masa, year)  — the folder processes the PREVIOUS month
MONTHS = [
    ("WORK JANUARY 2026", 12, 2025),
    ("WORK FEBRUARY 2026", 1, 2026),
    ("WORK MARCH 2026", 2, 2026),
    ("WORK APRIL 2026", 3, 2026),
    ("WORK MAY 2026", 4, 2026),
]


def pick(folder, ro, must, nots=()):
    base = f"{ROOT}/{folder}/PPN {ro}/"
    out = []
    for f in glob.glob(base + "*.xlsx"):
        b = os.path.basename(f).upper()
        if "~$" in b:
            continue
        if all(m.upper() in b for m in must) and not any(n.upper() in b for n in nots):
            out.append(f)
    return out


def sap_file(folder, ro):
    c = pick(folder, ro, ["SAP", ro], ["TEMPLATE", "EKSPOR"])
    return c[0] if c else None


def ct_file(folder, ro):
    c = pick(folder, ro, ["EKSPOR", ro], ["TEMPLATE"])
    return c[0] if c else None


def tpl_file(folder, ro):
    c = pick(folder, ro, ["TEMPLATE IMPOR", ro], ["GRUP", "KEDUA", "KETIGA"])
    return c[0] if c else None


def grup_file(folder, ro):
    c = pick(folder, ro, ["TEMPLATE IMPOR", ro, "GRUP"])
    return c[0] if c else None


def etb_file(folder):
    c = [f for f in glob.glob(f"{ROOT}/{folder}/*ETB*PPN WAPU*.xlsx") if "~$" not in f]
    return c[0] if c else None


def keyframe(df):
    d = df.copy()
    for c in ["NOMOR_FAKTUR", "NPWP_WP", "ID_TKU_WP", "NPWP_PENJUAL",
              "FIELD_TAMBAHAN_2", "FIELD_TAMBAHAN_1"]:
        if c in d:
            d[c] = d[c].astype(str).str.strip()
    for c in ["KONFIRMASI", "MASA_PAJAK", "TAHUN_PAJAK", "MASA_PENGKREDITAN", "TAHUN_PENGKREDITAN"]:
        if c in d:
            d[c] = pd.to_numeric(d[c], errors="coerce").astype("Int64").astype(str)
    return d.set_index("NOMOR_FAKTUR")


def template_check(res, tpl):
    exp = pd.read_excel(tpl, sheet_name="FM - Import", dtype=object)
    exp = exp[exp["FM"] == "FM"].reset_index(drop=True)
    E, G = keyframe(exp), keyframe(res.fm_import)
    common = sorted(set(E.index) & set(G.index))
    cols = [c for c in E.columns if c in G.columns]
    tot = mm = ctot = cmm = 0
    for fk in common:
        for c in cols:
            tot += 1
            bad = str(E.loc[fk, c]) != str(G.loc[fk, c])
            mm += bad
            if c != "FIELD_TAMBAHAN_2":
                ctot += 1; cmm += bad
    extra = len(set(G.index) - set(E.index)); miss = len(set(E.index) - set(G.index))
    return (f"rows {len(G)}/{len(E)} (+{extra}/-{miss}) cells {100*(tot-mm)/tot:.1f}% "
            f"tax-crit {100*(ctot-cmm)/ctot:.1f}%")


def recon_check(sap_path, ct, etb_path, cfg):
    sap = read_sap(sap_path)
    coretax = read_coretax(ct, "Faktur Masukan_1")
    etb = read_etb(etb_path, ro_name=cfg.ro_name)
    cbf, cba, cbs = build_cabang_index(sap_path)
    rekon, flag = build_rekon(coretax, sap, etb, cfg, cbf, cba, cbs)
    wb = openpyxl.load_workbook(sap_path, data_only=True)
    if "REKON" not in wb.sheetnames:
        wb.close(); return f"no REKON sheet (etb {len(etb)} ukers)", len(flag)
    rk = wb["REKON"]
    hdr = [str(c.value).strip().upper() if c.value else "" for c in rk[1]]
    col = lambda name: (hdr.index(name) + 1) if name in hdr else None
    etb_c, sel_c = col("ETB"), col("SELISIH")
    month_cols = [(hdr[j], j + 1) for j in range(len(hdr)) if hdr[j] in REKON_MONTHS]
    real = {}
    for i in range(2, rk.max_row + 1):
        u = rk.cell(i, 1).value
        if u in (None, ""): continue
        m = {}
        for nm, c in month_cols:
            v = rk.cell(i, c).value
            if v not in (None, 0):
                m[nm] = m.get(nm, 0) + v
        real[int(u)] = {"m": m,
                        "ETB": rk.cell(i, etb_c).value if etb_c else None,
                        "SEL": rk.cell(i, sel_c).value if sel_c else None}
    wb.close()
    gen = {int(r["KODE UKER"]): r for _, r in rekon.iterrows()}
    om = okm = oe = oke = osel = oksel = 0
    for u, info in real.items():
        g = gen.get(u)
        if g is None: continue
        for ml, v in info["m"].items():
            okm, om = (okm + 1, om) if (g.get(ml) is not None and abs(g[ml] - v) < 1) else (okm, om + 1)
        if info["ETB"] is not None:
            oke, oe = (oke + 1, oe) if (g.get("ETB") is not None and abs(g["ETB"] - info["ETB"]) < 1) else (oke, oe + 1)
        if info["SEL"] is not None:
            oksel, osel = (oksel + 1, osel) if (g.get("SELISIH") is not None and abs(g["SELISIH"] - info["SEL"]) < 1) else (oksel, osel + 1)
    return (f"ukers {len(gen)}/{len(real)} months {okm}ok/{om} ETB {oke}ok/{oe} "
            f"SEL {oksel}ok/{osel}"), len(flag)


def full_check(res, grup):
    exp = pd.read_excel(grup, sheet_name="FM - Import", dtype=object)
    exp.columns = [str(c).strip() for c in exp.columns]
    exp = exp[exp["FM"] == "FM"]
    g = {str(r["NOMOR_FAKTUR"]).strip(): r for _, r in res.fm_import_full.iterrows()}
    e = {str(r["NOMOR_FAKTUR"]).strip(): r for _, r in exp.iterrows()}
    com = set(g) & set(e)

    def num(col):
        ok = n = 0
        for fk in com:
            ev = e[fk].get(col)
            if pd.isna(ev): continue
            n += 1
            try: ok += abs(float(g[fk].get(col)) - float(ev)) < 1
            except: pass
        return f"{ok}/{n}"

    def txt(col):
        ok = n = 0
        for fk in com:
            ev = e[fk].get(col)
            if pd.isna(ev): continue
            n += 1
            if str(g[fk].get(col)).strip().upper() == str(ev).strip().upper(): ok += 1
        return f"{ok}/{n}"
    return f"Kode {num('Kode Uker')} Nama {txt('Nama Uker')} DPP {num('DPP')} SAPpajak {num('Jumlah Pajak di SAP')}"


print("=" * 96)
print("HISTORICAL VALIDATION — PPN WAPU")
print("=" * 96)
for folder, masa, year in MONTHS:
    for ro in ["PALEMBANG", "YOGYAKARTA"]:
        cfg = Config(ro_name=ro, masa=masa, tahun=year)
        sapf, ctf, tplf, grupf, etbf = (sap_file(folder, ro), ct_file(folder, ro),
                                        tpl_file(folder, ro), grup_file(folder, ro),
                                        etb_file(folder))
        tag = f"{folder.replace('WORK ','').replace(' 2026','')[:9]:9} {ro[:4]} (masa {masa}/{year})"
        if not (sapf and ctf):
            print(f"\n### {tag}  — SKIP (SAP/CT missing: sap={bool(sapf)} ct={bool(ctf)})")
            continue
        print(f"\n### {tag}")
        try:
            res = run(sapf, ctf, cfg, coretax_sheet="Faktur Masukan_1", etb_path=etbf)
            print("  TEMPLATE:", template_check(res, tplf) if tplf else "(no template file)")
            if etbf:
                r, nf = recon_check(sapf, ctf, etbf, cfg)
                print(f"  REKON   : {r} | flagged {nf}")
            print("  FULL    :", full_check(res, grupf) if grupf else "(no GRUP file)")
        except Exception as e:
            print(f"  ❌ ERROR: {type(e).__name__}: {e}")
            traceback.print_exc()
print("\n" + "=" * 96)
